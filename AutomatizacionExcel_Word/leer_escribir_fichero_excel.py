import openpyxl


print("LEER UN CONTENIDO DE UN FICHERO EXCEL")
libro = openpyxl.load_workbook('C:\\Users\\Victo\\Desktop\\IPA Automatizacion de Procesos Inteligentes con Python\\Distrito_poblacion.xlsx')

##Obtenemos las pestañanas del libro
print(libro.sheetnames)

# Accedemos a una hoja específica
pestaña = libro['Distrito_poblacion']

##obtenemos el valor de una celda
print(pestaña['A1'].value)

##Acceder a una celda concreta
print(pestaña.cell(row=2, column=1).value)

##Recorrer todos los valores de una columna )es este caso la primera columna)
for i in range(1,6):
    print(pestaña.cell(row=i, column=1).value)


##Recorrer todos los valores de una columna )es este caso la segunda columna)
for i in range(1,6):
    print(pestaña.cell(row=i, column=2).value)


print("-----------------------------------------")
print("ESCRIBIR UN CONTENIDO EN UN FICHERO EXCEL")

pestaña['C1'] = 'Ranking'

##Crear una nueva pestaña
pestaña2 = libro.create_sheet('Nueva_Pestaña')

##guardar los cambios
libro.save('C:\\Users\\Victo\\Desktop\\IPA Automatizacion de Procesos Inteligentes con Python\\Distrito_poblacion_modificado.xlsx')
